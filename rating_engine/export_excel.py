"""Genera un Excel di confronto tra il nostro rating e quello Sofascore."""

import sys, os
sys.path.insert(0, os.path.dirname(__file__))

import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule

from sofascore_loader import load_match_sofascore
from sofascore_dimensions import compute_d2 as sofa_d2, compute_d3 as sofa_d3
import d1_direct, d4_context, d5_role, formula


# ── palette colori ────────────────────────────────────────────────────────────
C_HEADER_BG   = "1F3C78"   # blu scuro
C_HEADER_FG   = "FFFFFF"
C_SUBHDR_BG   = "2E6DB4"
C_SUBHDR_FG   = "FFFFFF"
C_CANADA_BG   = "D6E8F3"
C_QATAR_BG    = "FFF0D6"
C_ALT_CANADA  = "EAF2FB"
C_ALT_QATAR   = "FFF8EC"
C_BORDER      = "B0C4DE"
C_GREEN       = "00B050"
C_RED         = "FF0000"
C_YELLOW      = "FFBF00"
C_NEUTRAL     = "808080"

THIN = Side(style="thin", color=C_BORDER)
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
THICK_LEFT = Border(left=Side(style="medium", color="4472C4"),
                    right=THIN, top=THIN, bottom=THIN)


def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color="000000", size=10):
    return Font(bold=bold, color=color, size=size, name="Calibri")

def _align(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def build_comparison(event_id: str) -> pd.DataFrame:
    """Pipeline completa, ritorna DataFrame con rating nostro + Sofascore."""
    print("Caricamento dati Sofascore...")
    events, positions = load_match_sofascore(event_id)

    print("Calcolo dimensioni...")
    d1 = d1_direct.compute(events, positions)
    d2 = sofa_d2(events)
    d3 = sofa_d3(events)
    d4 = d4_context.compute(events, positions)
    d5 = d5_role.compute(events, positions)
    result = formula.compute(events, positions, d1, d2, d3, d4, d5)

    # estrai rating Sofascore e team dai metadati eventi
    pid_sofa = {}
    pid_team = {}
    pid_stats = {}
    seen = set()
    for _, row in events.iterrows():
        pid = row.get("player_id")
        if pd.isna(pid) or int(pid) in seen:
            continue
        seen.add(int(pid))
        pid_sofa[int(pid)]  = row.get("_sofa_rating")
        pid_team[int(pid)]  = row.get("team", "")
        pid_stats[int(pid)] = row.get("_sofa_stats") or {}

    result["sofa_rating"]    = result["player_id"].map(pid_sofa)
    result["team"]           = result["player_id"].map(pid_team)
    result["delta"]          = (result["rating"] - result["sofa_rating"]).round(2)

    # aggiungi stat chiave per contesto
    def s(pid, key):
        return pid_stats.get(int(pid), {}).get(key, 0) or 0

    result["xG"]          = result["player_id"].apply(lambda p: round(s(p, "expectedGoals"), 2))
    result["xA"]          = result["player_id"].apply(lambda p: round(s(p, "expectedAssists"), 2))
    result["gol"]         = result["player_id"].apply(lambda p: int(s(p, "goals")))
    result["assist"]      = result["player_id"].apply(lambda p: int(s(p, "goalAssist")))
    result["tiri"]        = result["player_id"].apply(lambda p: int(s(p, "totalShots") or s(p, "onTargetScoringAttempt")))
    result["passaggi"]    = result["player_id"].apply(lambda p: int(s(p, "totalPass")))
    result["keyPass"]     = result["player_id"].apply(lambda p: int(s(p, "keyPass")))
    result["progressione"]= result["player_id"].apply(lambda p: round(s(p, "totalProgression"), 1))
    result["touches"]     = result["player_id"].apply(lambda p: int(s(p, "touches")))

    # colonna minuti: viene da formula.py (colonna "minutes")
    result = result.rename(columns={"minutes": "minuti"})
    result = result.sort_values(["team", "rating"], ascending=[True, False])
    return result


def write_excel(df: pd.DataFrame, path: str, home: str, away: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Confronto Rating"

    # ── TITOLO ──────────────────────────────────────────────────────────────
    ws.merge_cells("A1:P1")
    ws["A1"] = f"RATING CONFRONTO — {home.upper()} vs {away.upper()} | WC 2026"
    ws["A1"].font      = _font(bold=True, color=C_HEADER_FG, size=13)
    ws["A1"].fill      = _fill(C_HEADER_BG)
    ws["A1"].alignment = _align()
    ws.row_dimensions[1].height = 28

    # ── INTESTAZIONI ────────────────────────────────────────────────────────
    headers = [
        ("Giocatore",      18),
        ("Ruolo",          10),
        ("Min",             6),
        ("Gol",             5),
        ("Ast",             5),
        ("xG",              7),
        ("xA",              7),
        ("Tiri",            6),
        ("Key P",           6),
        ("Progr.",          8),
        ("D1\nAzioni",      8),
        ("D2\nxImpact",     8),
        ("D3\nNetwork",     8),
        ("D4\nContesto",    8),
        ("D5\nRole",        8),
        ("NOSTRO\nRATING", 10),
        ("SOFA\nRATING",   10),
        ("DELTA",           8),
    ]

    for col_idx, (hdr, width) in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx, value=hdr)
        cell.font      = _font(bold=True, color=C_SUBHDR_FG, size=9)
        cell.fill      = _fill(C_SUBHDR_BG)
        cell.alignment = _align(wrap=True)
        cell.border    = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[2].height = 30

    # ── DATI ────────────────────────────────────────────────────────────────
    home_teams = df[df["team"].str.lower().str.contains(home.lower(), na=False)]["player_id"].tolist()

    row_num = 3
    prev_team = None

    for _, r in df.iterrows():
        is_home   = r["team"].lower().find(home.lower()) >= 0
        is_alt    = (row_num % 2 == 0)
        bg_base   = C_CANADA_BG if is_home else C_QATAR_BG
        bg_alt    = C_ALT_CANADA if is_home else C_ALT_QATAR

        # separatore tra squadre
        if prev_team is not None and r["team"] != prev_team:
            ws.row_dimensions[row_num].height = 6
            for c in range(1, 19):
                ws.cell(row=row_num, column=c).fill = _fill("CCCCCC")
            row_num += 1

        prev_team = r["team"]
        bg = bg_alt if is_alt else bg_base
        fill = _fill(bg)

        values = [
            r["player_name"],
            r["role_group"],
            r["minuti"],
            r["gol"]      if r["gol"] else "",
            r["assist"]   if r["assist"] else "",
            r["xG"]       if r["xG"] else "",
            round(r["xA"], 2) if r["xA"] else "",
            r["tiri"]     if r["tiri"] else "",
            r["keyPass"]  if r["keyPass"] else "",
            r["progressione"],
            r["d1_norm"],
            r["d2_norm"],
            r["d3_norm"],
            r["d4_norm"],
            r["d5_norm"],
            r["rating"],
            r["sofa_rating"],
            r["delta"],
        ]

        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            cell.fill      = fill
            cell.border    = BORDER
            cell.alignment = _align(h="left" if col_idx == 1 else "center")
            cell.font      = _font(size=9)

        # evidenzia rating
        our_cell  = ws.cell(row=row_num, column=16)
        sofa_cell = ws.cell(row=row_num, column=17)
        delta_cell= ws.cell(row=row_num, column=18)

        our_cell.font  = _font(bold=True, size=10)
        sofa_cell.font = _font(bold=True, size=10)

        # colora delta
        delta = r["delta"]
        if pd.isna(delta):
            delta_color = C_NEUTRAL
        elif delta > 0.5:
            delta_color = C_GREEN
        elif delta < -0.5:
            delta_color = C_RED
        else:
            delta_color = C_YELLOW
        delta_cell.font = _font(bold=True, color=delta_color, size=10)

        ws.row_dimensions[row_num].height = 18
        row_num += 1

    # ── LEGENDA DELTA ────────────────────────────────────────────────────────
    row_num += 1
    ws.merge_cells(f"A{row_num}:D{row_num}")
    ws[f"A{row_num}"] = "DELTA = Nostro Rating − Sofascore Rating"
    ws[f"A{row_num}"].font = _font(bold=True, size=9, color="444444")

    row_num += 1
    for color, label in [(C_GREEN, "> +0.5  Sovrastimiamo vs Sofa"),
                          (C_YELLOW, "±0.5   In linea"),
                          (C_RED,   "< -0.5  Sottostimiamo vs Sofa")]:
        ws.merge_cells(f"A{row_num}:D{row_num}")
        ws[f"A{row_num}"] = label
        ws[f"A{row_num}"].font = _font(bold=True, size=9, color=color)
        row_num += 1

    # ── FREEZE e filtri ──────────────────────────────────────────────────────
    ws.freeze_panes = "A3"

    # ── FOGLIO 2: scatter dati per analisi ───────────────────────────────────
    ws2 = wb.create_sheet("Dati Raw")
    raw_cols = ["player_name", "team", "role_group", "minuti",
                "gol", "assist", "xG", "xA", "tiri", "keyPass", "progressione", "touches",
                "d1_norm", "d2_norm", "d3_norm", "d4_norm", "d5_norm", "composite",
                "rating", "sofa_rating", "delta"]

    for col_idx, col in enumerate(raw_cols, start=1):
        ws2.cell(row=1, column=col_idx, value=col).font = _font(bold=True, color="FFFFFF")
        ws2.cell(row=1, column=col_idx).fill = _fill(C_HEADER_BG)
        ws2.column_dimensions[get_column_letter(col_idx)].width = 14

    for row_idx, (_, r) in enumerate(df.iterrows(), start=2):
        for col_idx, col in enumerate(raw_cols, start=1):
            ws2.cell(row=row_idx, column=col_idx, value=r.get(col))

    wb.save(path)
    print(f"Excel salvato: {path}")


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("--event_id", default="15186798")
    parser.add_argument("--home",     default="Home")
    parser.add_argument("--away",     default="Away")
    parser.add_argument("--out",      default=None)
    args = parser.parse_args()

    out_path = args.out or os.path.join(
        os.path.dirname(__file__), "..",
        f"{args.home.lower().replace(' ', '_')}_vs_{args.away.lower().replace(' ', '_')}_rating.xlsx"
    )

    df = build_comparison(args.event_id)
    write_excel(df, out_path, args.home, args.away)

    print("\n" + "="*70)
    print(f"CONFRONTO RATING — {args.home} vs {args.away} (WC 2026)")
    print("="*70)
    cols = ["player_name", "team", "role_group", "minuti",
            "gol", "xG", "xA", "rating", "sofa_rating", "delta"]
    print(df[cols].to_string(index=False))
