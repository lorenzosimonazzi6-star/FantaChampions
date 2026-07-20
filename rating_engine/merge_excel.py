"""Genera tutti i rating WC 2026, accorpa in un unico Excel e salva in Downloads."""

import sys, os
sys.path.insert(0, os.path.dirname(__file__))

import glob
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from export_excel import build_comparison, write_excel

# ── partite da processare ─────────────────────────────────────────────────────
MATCHES = [
    ("15186798", "Canada",       "Qatar"),
    ("15186854", "Argentina",    "Algeria"),
    ("15186504", "Inghilterra",  "Croazia"),
    ("15186501", "Francia",      "Senegal"),
    ("15186850", "Brasile",      "Marocco"),
]

OUTPUT_PATH = os.path.join(os.path.expanduser("~"), "Downloads", "fantamondiale_rating_wc2026.xlsx")

BASE_DIR = os.path.join(os.path.dirname(__file__), "..")

# ── palette ───────────────────────────────────────────────────────────────────
C_HEADER_BG  = "1F3C78"
C_HEADER_FG  = "FFFFFF"
C_SUBHDR_BG  = "2E6DB4"
C_SUBHDR_FG  = "FFFFFF"
C_ROW_A      = "EAF2FB"
C_ROW_B      = "FFFFFF"
C_GREEN      = "00B050"
C_RED        = "FF0000"
C_YELLOW     = "FFBF00"
C_NEUTRAL    = "808080"

THIN   = Side(style="thin",   color="B0C4DE")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def _fill(c):  return PatternFill("solid", fgColor=c)
def _font(bold=False, color="000000", size=10):
    return Font(bold=bold, color=color, size=size, name="Calibri")
def _align(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


HEADERS = [
    ("Giocatore",     18),
    ("Ruolo",         10),
    ("Min",            6),
    ("Gol",            5),
    ("Ast",            5),
    ("xG",             7),
    ("xA",             7),
    ("Tiri",           6),
    ("Key P",          6),
    ("Progr.",         8),
    ("D1",             7),
    ("D2",             7),
    ("D3",             7),
    ("D4",             7),
    ("D5",             7),
    ("NOSTRO",        10),
    ("SOFA",          10),
    ("DELTA",          8),
]

COL_VALS = [
    "player_name", "role_group", "minuti",
    "gol", "assist", "xG", "xA", "tiri", "keyPass", "progressione",
    "d1_norm", "d2_norm", "d3_norm", "d4_norm", "d5_norm",
    "rating", "sofa_rating", "delta",
]


def _write_match_sheet(wb: openpyxl.Workbook, df: pd.DataFrame, home: str, away: str):
    sheet_name = f"{home[:10]} v {away[:10]}"
    ws = wb.create_sheet(title=sheet_name)

    # titolo
    ws.merge_cells(f"A1:{get_column_letter(len(HEADERS))}1")
    ws["A1"] = f"{home.upper()} vs {away.upper()}"
    ws["A1"].font      = _font(bold=True, color=C_HEADER_FG, size=12)
    ws["A1"].fill      = _fill(C_HEADER_BG)
    ws["A1"].alignment = _align()
    ws.row_dimensions[1].height = 24

    # intestazioni
    for ci, (hdr, width) in enumerate(HEADERS, 1):
        c = ws.cell(row=2, column=ci, value=hdr)
        c.font      = _font(bold=True, color=C_SUBHDR_FG, size=9)
        c.fill      = _fill(C_SUBHDR_BG)
        c.alignment = _align(wrap=True)
        c.border    = BORDER
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[2].height = 28

    # raggruppamento per squadra
    teams = df["team"].unique()
    home_team = next((t for t in teams if home.lower() in t.lower()), teams[0] if len(teams) else "")

    prev_team = None
    row_num   = 3
    alt       = False

    for _, r in df.iterrows():
        is_home = home.lower() in str(r["team"]).lower()
        if prev_team is not None and r["team"] != prev_team:
            # separatore tra squadre
            ws.row_dimensions[row_num].height = 5
            for ci in range(1, len(HEADERS) + 1):
                ws.cell(row=row_num, column=ci).fill = _fill("CCCCCC")
            row_num += 1
            alt = False
        prev_team = r["team"]

        bg = C_ROW_A if alt else C_ROW_B
        alt = not alt
        fill = _fill(bg)

        vals = [r.get(col, "") for col in COL_VALS]
        vals = [v if v != 0 or COL_VALS[i] in ("gol","assist","tiri","keyPass") else
                ("" if COL_VALS[i] in ("gol","assist","tiri","keyPass","xG","xA") else v)
                for i, v in enumerate(vals)]

        for ci, val in enumerate(vals, 1):
            cell = ws.cell(row=row_num, column=ci, value=val)
            cell.fill      = fill
            cell.border    = BORDER
            cell.alignment = _align(h="left" if ci == 1 else "center")
            cell.font      = _font(size=9)

        # rating in evidenza
        ws.cell(row=row_num, column=16).font = _font(bold=True, size=10)
        ws.cell(row=row_num, column=17).font = _font(bold=True, size=10)

        delta = r.get("delta")
        if pd.isna(delta) or delta is None:
            dc = C_NEUTRAL
        elif delta > 0.5:
            dc = C_GREEN
        elif delta < -0.5:
            dc = C_RED
        else:
            dc = C_YELLOW
        ws.cell(row=row_num, column=18).font = _font(bold=True, color=dc, size=10)

        ws.row_dimensions[row_num].height = 17
        row_num += 1

    ws.freeze_panes = "A3"


def _write_summary_sheet(wb: openpyxl.Workbook, all_dfs: list[tuple[str, str, pd.DataFrame]]):
    """Foglio riepilogo: tutti i giocatori, ordinati per rating."""
    ws = wb.create_sheet(title="Tutti i giocatori", index=0)

    extra_headers = [("Partita", 20)] + HEADERS
    ws.merge_cells(f"A1:{get_column_letter(len(extra_headers))}1")
    ws["A1"] = "RIEPILOGO GLOBALE — WC 2026"
    ws["A1"].font      = _font(bold=True, color=C_HEADER_FG, size=13)
    ws["A1"].fill      = _fill(C_HEADER_BG)
    ws["A1"].alignment = _align()
    ws.row_dimensions[1].height = 26

    for ci, (hdr, width) in enumerate(extra_headers, 1):
        c = ws.cell(row=2, column=ci, value=hdr)
        c.font      = _font(bold=True, color=C_SUBHDR_FG, size=9)
        c.fill      = _fill(C_SUBHDR_BG)
        c.alignment = _align(wrap=True)
        c.border    = BORDER
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[2].height = 28

    rows_all = []
    for home, away, df in all_dfs:
        for _, r in df.iterrows():
            rows_all.append({"_match": f"{home} vs {away}", **r.to_dict()})

    rows_all.sort(key=lambda x: x.get("rating", 0), reverse=True)

    alt = False
    for row_num, r in enumerate(rows_all, start=3):
        bg = C_ROW_A if alt else C_ROW_B
        alt = not alt
        fill = _fill(bg)

        all_vals = [r.get("_match", "")] + [r.get(col, "") for col in COL_VALS]
        for ci, val in enumerate(all_vals, 1):
            cell = ws.cell(row=row_num, column=ci, value=val)
            cell.fill      = fill
            cell.border    = BORDER
            cell.alignment = _align(h="left" if ci <= 2 else "center")
            cell.font      = _font(size=9)

        ws.cell(row=row_num, column=17).font = _font(bold=True, size=10)
        ws.cell(row=row_num, column=18).font = _font(bold=True, size=10)
        delta = r.get("delta")
        dc = C_NEUTRAL if (pd.isna(delta) if delta is not None else True) else (
             C_GREEN if delta > 0.5 else (C_RED if delta < -0.5 else C_YELLOW))
        ws.cell(row=row_num, column=19).font = _font(bold=True, color=dc, size=10)
        ws.row_dimensions[row_num].height = 17

    ws.freeze_panes = "A3"


if __name__ == "__main__":
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # rimuove foglio vuoto default

    all_dfs = []
    for event_id, home, away in MATCHES:
        print(f"\n{'='*60}")
        print(f"  {home} vs {away}  (eventId={event_id})")
        print(f"{'='*60}")
        df = build_comparison(event_id)
        all_dfs.append((home, away, df))
        _write_match_sheet(wb, df, home, away)

    _write_summary_sheet(wb, all_dfs)

    wb.save(OUTPUT_PATH)
    print(f"\nExcel salvato: {OUTPUT_PATH}")

    # elimina vecchi file singoli
    old_pattern = os.path.join(BASE_DIR, "*_rating.xlsx")
    deleted = []
    for f in glob.glob(old_pattern):
        os.remove(f)
        deleted.append(os.path.basename(f))

    if deleted:
        print(f"Eliminati {len(deleted)} file vecchi: {', '.join(deleted)}")
